#!/usr/bin/env python3
# 🕒 2026-02-17-00-00-00
# Police_Analytics_Dashboard/summons_etl_enhanced
# Author: R. A. Carucci
# Purpose: ETL pipeline for monthly E-ticket CSV processing.
# Mapping: Export (Officer Id) -> pad to 4 digits -> merge to Assignment Master on
#          PADDED_BADGE_NUMBER -> use Assignment Master 'Proposed 4-Digit Format' as
#          OFFICER_DISPLAY_NAME. Export name fields are never used for display.

import sys
import shutil
import pandas as pd
import numpy as np
from pathlib import Path
import logging
from datetime import datetime
import re
import json
from fuzzywuzzy import fuzz, process

# Ensure 06_Workspace_Management/scripts is on path for summons_backfill_merge and path_config
_master_scripts = Path(__file__).resolve().parents[2] / "06_Workspace_Management" / "scripts"
if _master_scripts.exists() and str(_master_scripts) not in sys.path:
    sys.path.insert(0, str(_master_scripts))

try:
    from path_config import get_onedrive_root
except ImportError:
    def get_onedrive_root():
        return Path(r"C:\Users\carucci_r\OneDrive - City of Hackensack")

# Manual overrides for temp assignments or badges not in Assignment Master
# Supports _condition: only apply when column contains substring (e.g. FIRE LANES)
# WG2 uses "SSOCC" to match Assignment Master display preference.
ASSIGNMENT_OVERRIDES = {
    # Feb 2026: Badge 2025 (Ramirez) - only FIRE LANES violations show as SSOCC; others stay Traffic Bureau.
    "2025": {
        "_condition": {"column": "Violation Description", "contains": "FIRE LANES"},
        "OFFICER_DISPLAY_NAME": "M. RAMIREZ #2025",
        "WG2": "SSOCC",
    },
    # Badge 0738 (Polson, CIV) - drone operator; not in Assignment Master; SSOCC for firezone summons.
    "0738": {
        "OFFICER_DISPLAY_NAME": "R. POLSON #0738",
        "WG2": "SSOCC",
    },
}

# ─── DFR Configuration ──────────────────────────────────────────────────────
DFR_CONFIG = {
    "dfr_badges": ["0738", "2025"],
    "target_workbook": "Shared Folder/Compstat/Contributions/Drone/dfr_directed_patrol_enforcement.xlsx",
    "target_sheet": "DFR Summons Log",
    "formula_columns": ["A", "G", "H", "I", "J", "P", "R"],
    "column_map": {
        "B": "Date",
        "C": "Time",
        "D": "Summons Number",
        "E": "Location",
        "F": "Statute",
        "K": "DFR Operator",
        "L": "Issuing Officer",
        "M": "Summons Status",
        "N": "DFR Unit ID",
        "O": "OCA",
        "Q": "Notes",
    },
    "status_map": {
        "ACTI": "Active",
        "VOID": "Void",
        "DISM": "Dismissed",
        "PAID": "Paid",
        "GUIL": "Guilty",
        "SUSP": "Suspended",
    },
    "summons_prefix": "E26",
}


def fix_data_types_quick(df):
    """Quick data type fix for save issues"""
    print(f"Fixing data types for {len(df)} records...")
    df = df.copy()

    # PADDED_BADGE_NUMBER must remain a string — removed from numeric list
    numeric_cols = ['FINE_AMOUNT', 'AMOUNT', 'DATA_QUALITY_SCORE']
    for col in numeric_cols:
        if col in df.columns:
            try:
                df[col] = pd.to_numeric(
                    df[col].astype(str).str.replace(r'[^\d.-]', '', regex=True),
                    errors='coerce'
                ).fillna(0)
                print(f"  Fixed numeric: {col}")
            except Exception as e:
                print(f"  Could not fix {col}: {e}")

    date_cols = ['VIOLATION_DATE', 'ISSUE_DATE', 'PROCESSED_TIMESTAMP']
    for col in date_cols:
        if col in df.columns:
            try:
                df[col] = pd.to_datetime(df[col], errors='coerce')
                print(f"  Fixed date: {col}")
            except Exception as e:
                print(f"  Could not fix {col}: {e}")

    string_cols = ['OFFICER_NAME_RAW', 'OFFICER_DISPLAY_NAME', 'DATA_SOURCE', 'DIVISION', 'BUREAU', 'FULL_NAME']
    for col in string_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).replace('nan', '')

    df = df.replace([np.inf, -np.inf], 0)
    print(f"Data type fixes complete")
    return df


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('summons_etl.log'),
        logging.StreamHandler()
    ]
)


class SummonsETLProcessor:
    def __init__(self, config_path="config.json"):
        self.base_path = Path(get_onedrive_root())
        self.source_data_path = self.base_path / "05_EXPORTS" / "_Summons" / "E_Ticket" / "2026" / "month"
        self.staging_path = self.base_path / "03_Staging" / "Summons"
        # Prefer 09_Reference/Personnel; fallback to 06_Workspace_Management root
        _am_ref = self.base_path / "09_Reference" / "Personnel" / "Assignment_Master_V2.csv"
        _am_ma = self.base_path / "06_Workspace_Management" / "Assignment_Master_V2.csv"
        self.assignment_master_path = _am_ref if _am_ref.exists() else _am_ma
        self.staging_path.mkdir(parents=True, exist_ok=True)
        self.assignment_master = self._load_assignment_master()

    # =========================================================================
    # FIX 1: _load_assignment_master — supports V2 (Proposed 4-Digit Format) and V3 (STANDARD_NAME)
    # =========================================================================
    def _load_assignment_master(self):
        """Load Assignment Master CSV.
        V2 schema: PADDED_BADGE_NUMBER + 'Proposed 4-Digit Format' (Jan 2026 and earlier)
        V3 schema: PADDED_BADGE_NUMBER + STANDARD_NAME (Feb 2026 onwards, PATROL BUREAU -> PATROL DIVISION)
        """
        try:
            df = pd.read_csv(self.assignment_master_path, dtype=str)
            logging.info(f"Assignment Master columns: {list(df.columns)}")

            if 'PADDED_BADGE_NUMBER' not in df.columns:
                logging.error("Assignment Master missing required column: PADDED_BADGE_NUMBER")
                return pd.DataFrame()

            # Accept either V2 ('Proposed 4-Digit Format') or V3 ('STANDARD_NAME') display name col
            if 'Proposed 4-Digit Format' in df.columns:
                display_col = 'Proposed 4-Digit Format'
                logging.info("Assignment Master: V2 schema detected (Proposed 4-Digit Format)")
            elif 'STANDARD_NAME' in df.columns:
                display_col = 'STANDARD_NAME'
                logging.info("Assignment Master: V3 schema detected (STANDARD_NAME / PATROL DIVISION)")
            else:
                logging.error("Assignment Master missing display name column (need 'Proposed 4-Digit Format' or 'STANDARD_NAME')")
                return pd.DataFrame()

            optional = {'WG2', 'LAST_NAME', 'TITLE'} & set(df.columns)
            df_clean = df[list({display_col, 'PADDED_BADGE_NUMBER'} | optional)].copy()

            df_clean = df_clean.rename(columns={display_col: 'OFFICER_DISPLAY_NAME'})
            df_clean['PADDED_BADGE_NUMBER'] = (
                df_clean['PADDED_BADGE_NUMBER'].astype(str).str.strip().str.zfill(4)
            )
            df_clean = df_clean.drop_duplicates(subset='PADDED_BADGE_NUMBER').reset_index(drop=True)

            logging.info(f"Loaded {len(df_clean)} officer records from Assignment Master")
            return df_clean

        except Exception as e:
            logging.error(f"Error loading Assignment Master: {e}")
            return pd.DataFrame()

    # =========================================================================
    # FIX 2: process_eticket_csv — NEW primary method for monthly e-ticket CSV
    #         Uses Officer Id as authoritative key; never parses name fields for identity
    # =========================================================================
    def process_eticket_csv(self, file_path):
        """
        Process monthly E-ticket CSV export (semicolon-delimited).
        Officer Id is the authoritative identity key. All name field variations
        are ignored for consolidation — records are enriched via Assignment Master
        join on PADDED_BADGE_NUMBER only.
        """
        file_path = Path(file_path)
        if not file_path.exists():
            logging.error(f"E-ticket file not found: {file_path}")
            return pd.DataFrame()
        logging.info(f"Loading e-ticket: {file_path.name}")

        try:
            if file_path.suffix.lower() == '.xlsx':
                df = pd.read_excel(file_path, dtype=str)
            else:
                df = pd.read_csv(file_path, sep=';', dtype=str, low_memory=False, on_bad_lines='skip')
                if 'Officer Id' not in df.columns and len(df.columns) < 5:
                    df = pd.read_csv(file_path, sep=',', dtype=str, low_memory=False, on_bad_lines='skip')
        except Exception as e:
            logging.error(f"Failed to read e-ticket file: {e}")
            return pd.DataFrame()

        logging.info(f"Loaded {len(df)} raw records | Columns: {list(df.columns)}")

        # ── Validate Officer Id ───────────────────────────────────────────────
        if 'Officer Id' not in df.columns:
            logging.error("'Officer Id' column not found. Cannot proceed.")
            return pd.DataFrame()

        df = df[df['Officer Id'].notna()].copy()
        df = df[df['Officer Id'].astype(str).str.strip().str.match(r'^\d+$')].copy()
        df = df[df['Officer Id'].astype(str).str.strip() != '0'].copy()
        logging.info(f"Records after Officer Id validation: {len(df)}")

        # ── Pad Officer Id → PADDED_BADGE_NUMBER ─────────────────────────────
        df['PADDED_BADGE_NUMBER'] = df['Officer Id'].astype(str).str.strip().str.zfill(4)

        # ── Preserve raw name for audit only (NOT used for identity) ──────────
        name_cols = ['Officer Last Name', 'Officer First Name', 'Officer Middle Initial']
        present = [c for c in name_cols if c in df.columns]
        if present:
            df['OFFICER_NAME_RAW'] = df[present].fillna('').agg(' '.join, axis=1).str.strip()
        else:
            df['OFFICER_NAME_RAW'] = ''

        # ── Standardize key columns ───────────────────────────────────────────
        if 'Case Type Code' in df.columns:
            df['Case Type Code'] = df['Case Type Code'].astype(str).str.strip().str.upper()
        if 'Statute' in df.columns:
            df = df.rename(columns={'Statute': 'STATUTE'})
        if 'Violation Date' in df.columns:
            df = df.rename(columns={'Violation Date': 'VIOLATION_DATE'})
        elif 'Issue Date' in df.columns:
            df = df.rename(columns={'Issue Date': 'VIOLATION_DATE'})

        df['DATA_SOURCE'] = f'ETICKET_{file_path.stem.upper()}'

        # ── Map to Assignment Master; display name = Proposed 4-Digit Format ───
        df = self._enrich_with_officer_data(df)
        df = self._categorize_violations(df)
        df = self._add_data_quality_metrics(df)
        df = self._flag_name_anomalies(df)

        logging.info(
            f"E-ticket processing complete: {len(df)} records, "
            f"{df['PADDED_BADGE_NUMBER'].nunique()} unique officers"
        )
        return df

    def _flag_name_anomalies(self, df):
        """QA flag: Officer Last Name contains spaces = full name stuffed in that field."""
        if 'Officer Last Name' in df.columns:
            df['NAME_FORMAT_ANOMALY'] = df['Officer Last Name'].astype(str).str.contains(r'\s', na=False)
        else:
            df['NAME_FORMAT_ANOMALY'] = False
        return df

    def _map_to_dfr_schema(self, df):
        """
        Transform enriched e-ticket DataFrame into DFR Summons Log schema.
        Only processes records whose PADDED_BADGE_NUMBER is in DFR_CONFIG['dfr_badges'].
        """
        dfr_badges = DFR_CONFIG['dfr_badges']
        dfr_df = df[df['PADDED_BADGE_NUMBER'].isin(dfr_badges)].copy()

        if dfr_df.empty:
            logging.info("No DFR records found in this batch")
            return pd.DataFrame()

        logging.info("DFR filter: %s records for badges %s", len(dfr_df), dfr_badges)

        mapped = pd.DataFrame()

        # B: Date
        if 'VIOLATION_DATE' in dfr_df.columns:
            mapped['Date'] = pd.to_datetime(dfr_df['VIOLATION_DATE'], errors='coerce').dt.date
        elif 'Issue Date' in dfr_df.columns:
            mapped['Date'] = pd.to_datetime(dfr_df['Issue Date'], errors='coerce').dt.date
        else:
            mapped['Date'] = pd.NaT

        # C: Time — 4-digit HHMM
        if 'Charge Time' in dfr_df.columns:
            mapped['Time'] = dfr_df['Charge Time'].astype(str).str.strip().str.zfill(4)
        else:
            mapped['Time'] = ''

        # D: Summons Number — strip prefix, keep 6 digits
        prefix = DFR_CONFIG['summons_prefix']
        if 'Ticket Number' in dfr_df.columns:
            mapped['Summons Number'] = (
                dfr_df['Ticket Number']
                .astype(str)
                .str.strip()
                .str.replace(prefix, '', regex=False)
                .str[-6:]
                .str.zfill(6)
            )
        else:
            mapped['Summons Number'] = ''

        # E: Location
        for loc_col in ('Offense Street Name', 'VIOLATION_LOCATION'):
            if loc_col in dfr_df.columns:
                mapped['Location'] = dfr_df[loc_col].astype(str).fillna('')
                break
        else:
            mapped['Location'] = ''

        # F: Statute
        stat_col = 'STATUTE' if 'STATUTE' in dfr_df.columns else 'Statute'
        mapped['Statute'] = dfr_df[stat_col].astype(str).fillna('') if stat_col in dfr_df.columns else ''

        # K, L: DFR Operator, Issuing Officer (from Assignment Master STANDARD_NAME)
        mapped['DFR Operator'] = dfr_df['OFFICER_DISPLAY_NAME'].astype(str).fillna('')
        mapped['Issuing Officer'] = mapped['DFR Operator']

        # M: Summons Status
        status_col = 'Case Status Code' if 'Case Status Code' in dfr_df.columns else 'Case Status'
        raw_status = dfr_df[status_col].astype(str).str.strip().str.upper() if status_col in dfr_df.columns else pd.Series('', index=dfr_df.index)
        status_map = DFR_CONFIG['status_map']
        mapped['Summons Status'] = raw_status.map(lambda x: status_map.get(x, x) if x else '')

        # N: DFR Unit ID (WG2 from Assignment Master)
        mapped['DFR Unit ID'] = dfr_df['WG2'].astype(str).fillna('') if 'WG2' in dfr_df.columns else ''

        # O: OCA (blank)
        mapped['OCA'] = ''

        # Q: Notes (ETL batch ID)
        mapped['Notes'] = f"ETL {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        return mapped

    def export_to_dfr_workbook(self, df):
        """
        Append DFR-mapped rows to dfr_directed_patrol_enforcement.xlsx.
        Deduplicates on Summons Number. Skips formula columns. On PermissionError, saves to .etl_temp.xlsx.
        """
        if df.empty:
            return

        try:
            from openpyxl import load_workbook
        except ImportError:
            logging.warning("openpyxl not available; skipping DFR workbook export")
            return

        target_path = self.base_path / DFR_CONFIG['target_workbook']
        if not target_path.exists():
            logging.warning("DFR target workbook not found: %s", target_path)
            return

        # Get existing Summons Numbers for dedup (column D = 4)
        existing_summons = set()
        try:
            wb = load_workbook(target_path, read_only=True, data_only=True)
            ws = wb[DFR_CONFIG['target_sheet']]
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                val = row[0].value
                if val is not None:
                    existing_summons.add(str(val).strip())
            wb.close()
        except Exception as e:
            logging.warning("Could not read existing DFR data for dedup: %s", e)

        # Filter out records already in workbook
        to_append = df[~df['Summons Number'].astype(str).str.strip().isin(existing_summons)].copy()
        if to_append.empty:
            logging.info("DFR export: 0 new rows (all %s already in workbook)", len(df))
            return

        try:
            wb = load_workbook(target_path)
        except PermissionError:
            temp_path = target_path.parent / ".etl_temp.xlsx"
            logging.warning("Target workbook open: saving to %s", temp_path)
            shutil.copy2(target_path, temp_path)
            wb = load_workbook(temp_path)
            target_path = temp_path

        ws = wb[DFR_CONFIG['target_sheet']]
        formula_cols = DFR_CONFIG['formula_columns']
        col_map = {c: i for i, c in enumerate("ABCDEFGHIJKLMNOPQR", 1)}
        next_row = ws.max_row + 1
        for _, r in to_append.iterrows():
            for col_letter, field in DFR_CONFIG['column_map'].items():
                if col_letter in formula_cols:
                    continue
                col_idx = col_map[col_letter]
                val = r.get(field, '')
                if pd.isna(val):
                    val = ''
                cell = ws.cell(row=next_row, column=col_idx, value=val)
                if col_letter in ("C", "D", "K", "L", "N"):
                    cell.number_format = '@'
            next_row += 1

        try:
            wb.save(target_path)
            logging.info("DFR export: appended %s rows to %s", len(to_append), target_path.name)
        except PermissionError:
            temp_path = target_path.parent / ".etl_temp.xlsx"
            wb.save(temp_path)
            logging.warning("Target workbook open: saved to %s", temp_path)
        finally:
            wb.close()

    # =========================================================================
    # FIX 3: _enrich_with_officer_data — badge-only join, produces OFFICER_DISPLAY_NAME
    # =========================================================================
    def _enrich_with_officer_data(self, df):
        """
        Map export to Assignment Master on PADDED_BADGE_NUMBER. OFFICER_DISPLAY_NAME
        is always from Assignment Master 'Proposed 4-Digit Format' (never from export).
        Also brings in WG2, TITLE, LAST_NAME. Unmatched badges get 'UNKNOWN - XXXX'.
        Optional: LAST_NAME_FUZZY_SCORE = fuzzy match export last name vs AM LAST_NAME.
        """
        if self.assignment_master.empty:
            logging.warning("Assignment Master not loaded — skipping enrichment")
            df['OFFICER_DISPLAY_NAME'] = 'UNKNOWN'
            df['WG2'] = ''
            df['OFFICER_MATCH_QUALITY'] = 'NO_MASTER'
            return df

        merged = df.merge(
            self.assignment_master,
            on='PADDED_BADGE_NUMBER',
            how='left',
            suffixes=('', '_AM')
        )

        unmatched_mask = merged['OFFICER_DISPLAY_NAME'].isna()
        logging.info(
            f"Badge join: {(~unmatched_mask).sum()} matched, {unmatched_mask.sum()} unmatched"
        )

        if unmatched_mask.sum() > 0:
            bad_badges = merged.loc[unmatched_mask, 'PADDED_BADGE_NUMBER'].unique()
            logging.warning(f"Unmatched badges (update Assignment Master): {list(bad_badges)}")
            merged.loc[unmatched_mask, 'OFFICER_DISPLAY_NAME'] = (
                'UNKNOWN - ' + merged.loc[unmatched_mask, 'PADDED_BADGE_NUMBER']
            )

        if 'WG2' not in merged.columns:
            merged['WG2'] = ''
        merged['WG2'] = merged['WG2'].fillna('')

        # Apply manual overrides (optionally conditional on violation type)
        for badge, overrides in ASSIGNMENT_OVERRIDES.items():
            mask = merged['PADDED_BADGE_NUMBER'] == badge
            cond = overrides.get('_condition')
            if cond and isinstance(cond, dict):
                col_name, substring = cond.get('column'), cond.get('contains', '')
                if col_name and substring:
                    for c in [col_name, 'VIOLATION_DESCRIPTION', 'Violation Description']:
                        if c in merged.columns:
                            mask = mask & merged[c].astype(str).str.upper().str.contains(substring.upper(), na=False)
                            break
                overrides = {k: v for k, v in overrides.items() if k != '_condition'}
            if mask.any():
                logging.info(f"Applying assignment override for badge {badge} ({mask.sum()} records)")
                for col, value in overrides.items():
                    if col in merged.columns:
                        merged.loc[mask, col] = value

        # Normalize WG2: display "SSOCC" instead of "SAFE STREETS OPERATIONS CONTROL CENTER"
        if "WG2" in merged.columns:
            merged["WG2"] = merged["WG2"].replace(
                "SAFE STREETS OPERATIONS CONTROL CENTER", "SSOCC"
            )

        merged['OFFICER_MATCH_QUALITY'] = np.where(
            merged['OFFICER_DISPLAY_NAME'].str.startswith('UNKNOWN', na=False),
            'NO_MATCH',
            'DIRECT_BADGE'
        )

        # Fuzzy match: export "Officer Last Name" vs Assignment Master "LAST_NAME", verified by badge
        merged['LAST_NAME_FUZZY_SCORE'] = np.nan
        export_ln = 'Officer Last Name'
        am_ln = 'LAST_NAME'
        if export_ln in merged.columns and am_ln in merged.columns:
            matched_mask = merged['OFFICER_MATCH_QUALITY'] == 'DIRECT_BADGE'

            def _fuzz_score(row):
                a = str(row[export_ln] or '').strip().upper()
                b = str(row[am_ln] or '').strip().upper()
                if not a or not b:
                    return np.nan
                return fuzz.token_set_ratio(a, b)

            merged.loc[matched_mask, 'LAST_NAME_FUZZY_SCORE'] = merged.loc[matched_mask].apply(_fuzz_score, axis=1)
            low = (merged['LAST_NAME_FUZZY_SCORE'].notna()) & (merged['LAST_NAME_FUZZY_SCORE'] < 70)
            if low.any():
                n_low = int(low.sum())
                logging.warning(
                    "Last-name fuzzy match: %s rows with score < 70 (export vs Assignment Master by badge)",
                    n_low,
                )
        return merged

    # =========================================================================
    # UNCHANGED METHODS
    # =========================================================================
    def process_summons_master_file(self, file_path, sheet_name='ENHANCED_BADGE'):
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            logging.info(f"Loaded {len(df)} records from {sheet_name} sheet")
            actual_data_start = self._find_data_start_row(df)
            if actual_data_start > 0:
                df = pd.read_excel(file_path, sheet_name=sheet_name,
                                   header=actual_data_start, skiprows=range(actual_data_start))
            df = self._standardize_summons_columns(df)
            df = self._extract_badge_numbers(df)
            df_enriched = self._enrich_with_officer_data(df)
            df_final = self._categorize_violations(df_enriched)
            df_final = self._add_data_quality_metrics(df_final)
            return df_final
        except Exception as e:
            logging.error(f"Error processing summons file: {e}")
            return pd.DataFrame()

    def _find_data_start_row(self, df):
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            if any(isinstance(val, str) and any(p in str(val).upper() for p in
                  ['BADGE', 'OFFICER', 'VIOLATION', 'STATUTE', 'DATE']) for val in row):
                return i
        return 0

    def _standardize_summons_columns(self, df):
        column_mapping = {
            'Badge': 'BADGE_RAW', 'Badge Number': 'BADGE_RAW',
            'Officer': 'OFFICER_NAME_RAW', 'Officer Name': 'OFFICER_NAME_RAW',
            'Violation': 'VIOLATION_CODE', 'Statute': 'STATUTE',
            'Date': 'VIOLATION_DATE', 'Issue Date': 'VIOLATION_DATE',
            'Amount': 'FINE_AMOUNT', 'Fine': 'FINE_AMOUNT',
            'Location': 'VIOLATION_LOCATION'
        }
        for old_col, new_col in column_mapping.items():
            if old_col in df.columns:
                df = df.rename(columns={old_col: new_col})
        return df

    def _extract_badge_numbers(self, df):
        """Legacy badge extraction for non-eticket sources"""
        badge_columns = [col for col in df.columns if 'BADGE' in col.upper()]
        if not badge_columns:
            if 'OFFICER_NAME_RAW' in df.columns:
                df['BADGE_EXTRACTED'] = df['OFFICER_NAME_RAW'].apply(self._extract_badge_from_name)
            else:
                df['BADGE_EXTRACTED'] = None
        else:
            df['BADGE_EXTRACTED'] = df[badge_columns[0]]
        df['PADDED_BADGE_NUMBER'] = df['BADGE_EXTRACTED'].apply(
            lambda x: str(x).zfill(4) if pd.notna(x) and str(x).strip().isdigit() else None
        )
        return df

    def _extract_badge_from_name(self, officer_name):
        if pd.isna(officer_name):
            return None
        for pattern in [r'#(\d{3,4})', r'\s(\d{3,4})$', r'(\d{4})']:
            match = re.search(pattern, str(officer_name))
            if match:
                return match.group(1)
        return None

    def _categorize_violations(self, df):
        def categorize_violation(statute):
            if pd.isna(statute):
                return 'UNKNOWN'
            s = str(statute).upper()
            if '39:' in s or 'TITLE 39' in s:
                return 'TRAFFIC_TITLE39'
            elif any(o in s for o in ['ORD', 'ORDINANCE', 'MUNICIPAL']):
                return 'MUNICIPAL_ORDINANCE'
            return 'OTHER'

        df['VIOLATION_CATEGORY'] = df.get('STATUTE', pd.Series()).apply(categorize_violation)

        def classify_type(row):
            # Use only export Case Type Code (M=Moving, P=Parking, C=Other); no statute logic
            raw_type = str(row.get('Case Type Code', '')).strip().upper()
            if raw_type in ['M', 'P', 'C']:
                return raw_type
            return "P"

        df['TYPE'] = df.apply(classify_type, axis=1)

        if 'VIOLATION_DATE' in df.columns:
            df['ISSUE_DATE'] = pd.to_datetime(df['VIOLATION_DATE'], errors='coerce')
            df['Month_Year'] = df['ISSUE_DATE'].dt.strftime('%m-%y')
            df['Year'] = df['ISSUE_DATE'].dt.year
            df['Month'] = df['ISSUE_DATE'].dt.month
            df['YearMonthKey'] = (df['Year'] * 100 + df['Month']).fillna(0).astype(int)

        return df

    def _add_data_quality_metrics(self, df):
        df['DATA_QUALITY_SCORE'] = 0
        df['DATA_QUALITY_ISSUES'] = ''
        if 'PADDED_BADGE_NUMBER' in df.columns:
            df.loc[df['PADDED_BADGE_NUMBER'].notna(), 'DATA_QUALITY_SCORE'] += 25
            df.loc[df['PADDED_BADGE_NUMBER'].isna(), 'DATA_QUALITY_ISSUES'] += 'MISSING_BADGE; '
        if 'OFFICER_MATCH_QUALITY' in df.columns:
            df.loc[df['OFFICER_MATCH_QUALITY'] == 'DIRECT_BADGE', 'DATA_QUALITY_SCORE'] += 25
            df.loc[df['OFFICER_MATCH_QUALITY'] == 'NO_MATCH', 'DATA_QUALITY_ISSUES'] += 'NO_OFFICER_MATCH; '
        if 'VIOLATION_DATE' in df.columns:
            df.loc[df['VIOLATION_DATE'].notna(), 'DATA_QUALITY_SCORE'] += 25
            df.loc[df['VIOLATION_DATE'].isna(), 'DATA_QUALITY_ISSUES'] += 'MISSING_DATE; '
        if 'STATUTE' in df.columns:
            df.loc[df['STATUTE'].notna(), 'DATA_QUALITY_SCORE'] += 25
            df.loc[df['STATUTE'].isna(), 'DATA_QUALITY_ISSUES'] += 'MISSING_STATUTE; '
        df['DATA_QUALITY_ISSUES'] = df['DATA_QUALITY_ISSUES'].str.rstrip('; ')
        return df

    def process_ats_report(self, file_path):
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            logging.info(f"Loaded {len(df)} ATS records")
            if len(df.columns) >= 6:
                df.columns = ['BADGE_RAW', 'OFFICER_NAME_RAW', 'CITATION_REF', 'CITATION_NUMBER',
                              'VIOLATION_DATE', 'STATUTE'] + list(df.columns[6:])
            df['PADDED_BADGE_NUMBER'] = df['BADGE_RAW'].astype(str).str.zfill(4)
            df['DATA_SOURCE'] = 'ATS_AUTOMATED'
            df['VIOLATION_CATEGORY'] = 'TRAFFIC_AUTOMATED'
            return self._enrich_with_officer_data(df)
        except Exception as e:
            logging.error(f"Error processing ATS report: {e}")
            return pd.DataFrame()

    def create_unified_summons_dataset(self, eticket_files=None, summons_files=None, ats_files=None):
        all_datasets = []
        if eticket_files:
            for fp in eticket_files:
                df = self.process_eticket_csv(fp)
                if not df.empty:
                    all_datasets.append(df)
        if summons_files:
            for fp in summons_files:
                df = self.process_summons_master_file(fp)
                if not df.empty:
                    df['DATA_SOURCE'] = f'MANUAL_SUMMONS_{Path(fp).stem}'
                    all_datasets.append(df)
        if ats_files:
            for fp in ats_files:
                df = self.process_ats_report(fp)
                if not df.empty:
                    all_datasets.append(df)

        if not all_datasets:
            logging.warning("No valid datasets found")
            return pd.DataFrame()

        unified_df = pd.concat(all_datasets, ignore_index=True, sort=False)
        unified_df['PROCESSED_TIMESTAMP'] = datetime.now()
        unified_df['ETL_VERSION'] = '2.1'
        self._generate_processing_summary(unified_df)
        return unified_df

    def _generate_processing_summary(self, df):
        summary = {
            'total_records': len(df),
            'date_range': f"{df['VIOLATION_DATE'].min()} to {df['VIOLATION_DATE'].max()}"
                          if 'VIOLATION_DATE' in df.columns else 'Unknown',
            'unique_officers': df['PADDED_BADGE_NUMBER'].nunique() if 'PADDED_BADGE_NUMBER' in df.columns else 0,
            'match_quality_distribution': df['OFFICER_MATCH_QUALITY'].value_counts().to_dict()
                                          if 'OFFICER_MATCH_QUALITY' in df.columns else {},
            'type_distribution': df['TYPE'].value_counts().to_dict() if 'TYPE' in df.columns else {},
            'name_anomaly_count': int(df['NAME_FORMAT_ANOMALY'].sum()) if 'NAME_FORMAT_ANOMALY' in df.columns else 0,
            'data_quality_avg': df['DATA_QUALITY_SCORE'].mean() if 'DATA_QUALITY_SCORE' in df.columns else 0
        }
        logging.info("=== PROCESSING SUMMARY ===")
        for k, v in summary.items():
            logging.info(f"{k}: {v}")
        summary_path = self.staging_path / f"processing_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(summary_path, 'w') as f:
            json.dump(summary, f, indent=2, default=str)

    def save_to_staging(self, df, filename_prefix="summons_powerbi"):
        """Save to timestamped file AND overwrite summons_powerbi_latest.xlsx"""
        df = fix_data_types_quick(df)

        # Column order — Power BI priority columns first
        pbi_priority = [
            'PADDED_BADGE_NUMBER', 'OFFICER_DISPLAY_NAME', 'WG2', 'TITLE', 'TYPE',
            'YearMonthKey', 'Month_Year', 'Year', 'Month',
            'VIOLATION_DATE', 'ISSUE_DATE', 'STATUTE', 'VIOLATION_CATEGORY',
            'OFFICER_NAME_RAW', 'NAME_FORMAT_ANOMALY',
            'OFFICER_MATCH_QUALITY', 'DATA_QUALITY_SCORE', 'DATA_QUALITY_ISSUES',
            'DATA_SOURCE', 'PROCESSED_TIMESTAMP', 'ETL_VERSION'
        ]
        cols_ordered = [c for c in pbi_priority if c in df.columns]
        remaining = [c for c in df.columns if c not in cols_ordered]
        df_out = df[cols_ordered + remaining]

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        ts_path = self.staging_path / f"{filename_prefix}_{timestamp}.xlsx"
        latest_path = self.staging_path / "summons_powerbi_latest.xlsx"

        # Write to timestamped file first (avoids corruption from overwriting open file)
        with pd.ExcelWriter(ts_path, engine='openpyxl') as writer:
            df_out.to_excel(writer, sheet_name='Summons_Data', index=False)
            worksheet = writer.sheets['Summons_Data']
            badge_col_idx = df_out.columns.get_loc('PADDED_BADGE_NUMBER') + 1
            for row in range(2, len(df_out) + 2):
                cell = worksheet.cell(row=row, column=badge_col_idx)
                cell.number_format = '@'
        logging.info(f"Saved: {ts_path}")

        # Copy to latest (avoids second Excel write; reduces OneDrive corruption risk)
        if ts_path.exists() and ts_path.stat().st_size > 1_000_000:
            if latest_path.exists():
                latest_path.unlink()
            shutil.copy2(ts_path, latest_path)
            logging.info(f"Copied to: {latest_path}")
        else:
            logging.warning("Timestamped file too small or missing; writing latest directly")
            with pd.ExcelWriter(latest_path, engine='openpyxl') as writer:
                df_out.to_excel(writer, sheet_name='Summons_Data', index=False)
                worksheet = writer.sheets['Summons_Data']
                badge_col_idx = df_out.columns.get_loc('PADDED_BADGE_NUMBER') + 1
                for row in range(2, len(df_out) + 2):
                    cell = worksheet.cell(row=row, column=badge_col_idx)
                    cell.number_format = '@'
            logging.info(f"Saved: {latest_path}")

        self._validate_record_counts(df_out)
        return self.staging_path / "summons_powerbi_latest.xlsx"

    def _validate_record_counts(self, df):
        """Post-save QA: log counts by badge+type to validate against raw CSV pivot"""
        if 'PADDED_BADGE_NUMBER' not in df.columns:
            return
        logging.info("=== QA VALIDATION: counts by badge+type ===")
        summary = (
            df.groupby(['PADDED_BADGE_NUMBER', 'OFFICER_DISPLAY_NAME', 'TYPE'])
            .size().reset_index(name='COUNT')
            .sort_values(['PADDED_BADGE_NUMBER', 'TYPE'])
        )
        logging.info(f"\n{summary.to_string(index=False)}")
        unknown = df[df['OFFICER_DISPLAY_NAME'].str.startswith('UNKNOWN', na=False)]
        if not unknown.empty:
            logging.warning("%s records with no Assignment Master match:", len(unknown))
            logging.warning(unknown['PADDED_BADGE_NUMBER'].value_counts().to_string())


# =============================================================================
# MAIN
# =============================================================================
def main():
    logging.info("Starting Summons ETL Process v2.1")
    processor = SummonsETLProcessor()

    # Load full 13-month window of e-ticket exports (report month = previous month)
    eticket_base = processor.base_path / "05_EXPORTS" / "_Summons" / "E_Ticket"
    report_end = datetime.now().replace(day=1) - pd.Timedelta(days=1)  # last day of prev month
    report_start = (report_end.replace(day=1) - pd.DateOffset(months=12))  # first day, 12 months back
    eticket_files = []
    search_folders = [
        lambda y, m: eticket_base / str(y) / "month",   # 2025/month/, 2026/month/
        lambda y, m: eticket_base / str(y),             # 2025/, 2026/
        lambda y, m: eticket_base,                     # root (e.g. 2025_07 at root)
    ]
    for d in pd.date_range(start=report_start, end=report_end.replace(day=1), freq="MS"):
        yyyy, mm = d.year, d.month
        name = f"{yyyy}_{mm:02d}_eticket_export.csv"
        for folder_fn in search_folders:
            fp = folder_fn(yyyy, mm) / name
            if fp.exists():
                eticket_files.append(fp)
                break
    eticket_files = list(dict.fromkeys(eticket_files))  # dedupe, preserve order
    if len(eticket_files) < 5:
        logging.warning(
            "Only %s e-ticket files found for 13-month window; backfill will supply historical months",
            len(eticket_files),
        )
    if not eticket_files:
        logging.error(
            "No e-ticket files found for 13-month window %s through %s",
            report_start.strftime("%Y-%m"),
            report_end.strftime("%Y-%m"),
        )
        return

    logging.info(
        "Loading %s e-ticket files (13-month window): %s",
        len(eticket_files),
        [f.name for f in eticket_files],
    )

    unified_data = processor.create_unified_summons_dataset(
        eticket_files=eticket_files
        # ats_files=[...]  # add when available
    )

    # Add TICKET_COUNT=1 for ticket-level rows so backfill merge can add aggregated rows
    if not unified_data.empty and "TICKET_COUNT" not in unified_data.columns:
        unified_data["TICKET_COUNT"] = 1

    # Merge backfill for gap months (03-25, 07-25, 10-25, 11-25) into Department-Wide Summons
    try:
        from summons_backfill_merge import merge_missing_summons_months
        unified_data = merge_missing_summons_months(unified_data)
    except ImportError as e:
        logging.warning("summons_backfill_merge not available; skipping backfill: %s", e)

    if not unified_data.empty:
        output_file = processor.save_to_staging(unified_data)
        logging.info("ETL complete. Output: %s", output_file)
        # DFR: filter, map, append to dfr_directed_patrol_enforcement.xlsx
        dfr_mapped = processor._map_to_dfr_schema(unified_data)
        if not dfr_mapped.empty:
            processor.export_to_dfr_workbook(dfr_mapped)
    else:
        logging.error("ETL failed - no data processed")


if __name__ == "__main__":
    main()



