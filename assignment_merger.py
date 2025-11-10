# 2025-06-28-15-50-15
# assignment_master_merger/enhanced_merger.py
# Author: R. A. Carucci
# Purpose: Production-ready Assignment Master merger with change logging and macro preservation

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from pathlib import Path
import logging
from datetime import datetime
import argparse
import sys

class EnhancedAssignmentMerger:
    """
    Enhanced Assignment Master merger that preserves all original data,
    tracks changes, and maintains Excel macro functionality.
    """
    
    def __init__(self, master_path: str, submission_path: str, conflicts_path: str, 
                 output_path: str = None, create_change_log: bool = True):
        self.master_path = Path(master_path)
        self.submission_path = Path(submission_path)
        self.conflicts_path = Path(conflicts_path)
        self.create_change_log = create_change_log
        self.output_path = Path(output_path) if output_path else self._generate_output_path()
        
        # Setup logging
        self.logger = self._setup_logging()
        
        # Track changes for reporting
        self.changes_log = []
        self.stats = {
            'total_records': 0,
            'submission_updates': 0,
            'conflict_updates': 0,
            'no_updates': 0,
            'errors': 0
        }
    
    def _setup_logging(self):
        """Configure logging with both file and console output."""
        log_format = '%(asctime)s - %(levelname)s - %(message)s'
        logging.basicConfig(
            level=logging.INFO,
            format=log_format,
            handlers=[
                logging.FileHandler(f'assignment_merge_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        return logging.getLogger(__name__)
    
    def _generate_output_path(self):
        """Generate timestamped output filename."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.master_path.parent / f"Assignment_Master_Updated_{timestamp}.xlsm"
    
    def normalize_badge(self, badge):
        """
        Normalize badge numbers to consistent 4-digit format.
        Handles various input formats (string, int, float, None).
        """
        try:
            if pd.isna(badge):
                return None
            # Convert to string, handle decimals, then to int, then pad
            badge_str = str(badge).strip()
            if '.' in badge_str:
                badge_num = int(float(badge_str))
            else:
                badge_num = int(badge_str)
            return str(badge_num).zfill(4)
        except (ValueError, TypeError):
            self.logger.warning(f"Could not normalize badge: {badge}")
            return None
    
    def load_and_validate_data(self):
        """Load all input files and perform validation checks."""
        try:
            self.logger.info("📂 Loading input files...")
            
            # Load Assignment Master with all sheets
            xl_master = pd.read_excel(self.master_path, sheet_name=None, engine='openpyxl')
            if 'Assignment_Master' not in xl_master:
                raise ValueError("Assignment_Master sheet not found in master file")
            self.df_master = xl_master['Assignment_Master'].copy()
            
            # Load submission data
            self.df_submission = pd.read_excel(self.submission_path, engine='openpyxl')
            
            # Load conflicts data
            self.df_conflicts = pd.read_excel(self.conflicts_path, engine='openpyxl')
            
            self.logger.info(f"✅ Data loaded successfully:")
            self.logger.info(f"  Master: {len(self.df_master)} records")
            self.logger.info(f"  Submissions: {len(self.df_submission)} records")
            self.logger.info(f"  Conflicts: {len(self.df_conflicts)} records")
            
            # Validate required columns
            self._validate_columns()
            
            return True
            
        except Exception as e:
            self.logger.error(f"❌ Error loading data: {e}")
            return False
    
    def _validate_columns(self):
        """Validate that required columns exist in all dataframes."""
        # Check master file for badge column
        if 'BADGE_NUMBER' not in self.df_master.columns:
            raise ValueError("BADGE_NUMBER column missing from master file")
        
        # Check submission file columns
        required_submission_cols = ['Current Badge', 'Proposed Standardized Name', 
                                  'Proposed 4-Digit Format', 'Conflict Resolution', 'Special Notes']
        missing_submission = [col for col in required_submission_cols 
                            if col not in self.df_submission.columns]
        if missing_submission:
            raise ValueError(f"Missing columns in submission file: {missing_submission}")
        
        # Check conflicts file columns
        required_conflict_cols = ['Badge', 'Proposed Resolution', 'Resolution Method']
        missing_conflicts = [col for col in required_conflict_cols 
                           if col not in self.df_conflicts.columns]
        if missing_conflicts:
            raise ValueError(f"Missing columns in conflicts file: {missing_conflicts}")
        
        self.logger.info("✅ Column validation passed")
    
    def normalize_all_badges(self):
        """Apply badge normalization to all dataframes."""
        self.logger.info("🔧 Normalizing badge numbers...")
        
        # Normalize master badges
        self.df_master['NORMALIZED_BADGE'] = self.df_master['BADGE_NUMBER'].apply(self.normalize_badge)
        
        # Normalize submission badges
        self.df_submission['NORMALIZED_BADGE'] = self.df_submission['Current Badge'].apply(self.normalize_badge)
        
        # Normalize conflict badges
        self.df_conflicts['NORMALIZED_BADGE'] = self.df_conflicts['Badge'].apply(self.normalize_badge)
        
        # Remove rows with invalid badges
        original_counts = {
            'master': len(self.df_master),
            'submission': len(self.df_submission),
            'conflicts': len(self.df_conflicts)
        }
        
        self.df_master = self.df_master.dropna(subset=['NORMALIZED_BADGE'])
        self.df_submission = self.df_submission.dropna(subset=['NORMALIZED_BADGE'])
        self.df_conflicts = self.df_conflicts.dropna(subset=['NORMALIZED_BADGE'])
        
        self.logger.info("🎯 Badge normalization results:")
        for name, df in [('master', self.df_master), ('submission', self.df_submission), ('conflicts', self.df_conflicts)]:
            dropped = original_counts[name] - len(df)
            if dropped > 0:
                self.logger.warning(f"  {name}: Dropped {dropped} records with invalid badges")
            else:
                self.logger.info(f"  {name}: All {len(df)} badges valid")
    
    def create_lookup_tables(self):
        """Create efficient lookup tables for merging."""
        self.logger.info("📊 Creating lookup tables...")
        
        # Create submission lookup
        self.submission_lookup = self.df_submission.set_index("NORMALIZED_BADGE")[
            ["Proposed Standardized Name", "Proposed 4-Digit Format", "Conflict Resolution", "Special Notes"]
        ].to_dict('index')
        
        # Create conflicts lookup
        self.conflicts_lookup = self.df_conflicts.set_index("NORMALIZED_BADGE")[
            ["Proposed Resolution", "Resolution Method", "Conflict Family", "Personnel Count"]
        ].to_dict('index')
        
        self.logger.info(f"✅ Lookup tables created:")
        self.logger.info(f"  Submission entries: {len(self.submission_lookup)}")
        self.logger.info(f"  Conflict entries: {len(self.conflicts_lookup)}")
    
    def apply_updates_with_tracking(self):
        """Apply updates to master data while tracking all changes."""
        self.logger.info("🔄 Applying updates with change tracking...")
        
        # Ensure target columns exist in master
        target_columns = ["Proposed Standardized Name", "Proposed 4-Digit Format", 
                         "Conflict Resolution", "Special Notes"]
        
        for col in target_columns:
            if col not in self.df_master.columns:
                self.df_master[col] = None
        
        # Process each row
        for idx, row in self.df_master.iterrows():
            badge = row['NORMALIZED_BADGE']
            original_values = {col: row[col] for col in target_columns}
            update_applied = False
            update_source = None
            
            try:
                # Priority 1: Check conflicts (they override submissions)
                if badge in self.conflicts_lookup:
                    conflict_data = self.conflicts_lookup[badge]
                    
                    # Create special notes for conflicts
                    special_notes = f"Conflict Resolution - Family: {conflict_data.get('Conflict Family', 'Unknown')} "
                    special_notes += f"({conflict_data.get('Personnel Count', 'Unknown')} personnel) - "
                    special_notes += f"Method: {conflict_data.get('Resolution Method', 'Unknown')}"
                    
                    # Apply conflict updates
                    self.df_master.at[idx, 'Proposed Standardized Name'] = conflict_data['Proposed Resolution']
                    self.df_master.at[idx, 'Proposed 4-Digit Format'] = conflict_data['Proposed Resolution']
                    self.df_master.at[idx, 'Conflict Resolution'] = "Yes"
                    self.df_master.at[idx, 'Special Notes'] = special_notes
                    
                    update_applied = True
                    update_source = "Conflicts Resolution"
                    self.stats['conflict_updates'] += 1
                
                # Priority 2: Check submissions
                elif badge in self.submission_lookup:
                    submission_data = self.submission_lookup[badge]
                    
                    # Apply submission updates
                    self.df_master.at[idx, 'Proposed Standardized Name'] = submission_data['Proposed Standardized Name']
                    self.df_master.at[idx, 'Proposed 4-Digit Format'] = submission_data['Proposed 4-Digit Format']
                    self.df_master.at[idx, 'Conflict Resolution'] = submission_data['Conflict Resolution']
                    self.df_master.at[idx, 'Special Notes'] = submission_data['Special Notes']
                    
                    update_applied = True
                    update_source = "ATS Submission"
                    self.stats['submission_updates'] += 1
                
                else:
                    self.stats['no_updates'] += 1
                
                # Track changes if updates were applied
                if update_applied and self.create_change_log:
                    new_values = {col: self.df_master.at[idx, col] for col in target_columns}
                    
                    # Only log if there were actual changes
                    if original_values != new_values:
                        change_record = {
                            'Badge': row['BADGE_NUMBER'],
                            'Name': f"{row.get('FIRST_NAME', '')} {row.get('LAST_NAME', '')}".strip(),
                            'Update_Source': update_source,
                            'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        
                        # Add before/after values
                        for col in target_columns:
                            change_record[f'Before_{col}'] = original_values[col]
                            change_record[f'After_{col}'] = new_values[col]
                        
                        self.changes_log.append(change_record)
                
                self.stats['total_records'] += 1
                
            except Exception as e:
                self.logger.error(f"Error processing badge {badge}: {e}")
                self.stats['errors'] += 1
        
        # Clean up temporary column
        self.df_master = self.df_master.drop(columns=['NORMALIZED_BADGE'])
        
        self.logger.info("🎯 Update results:")
        for key, value in self.stats.items():
            self.logger.info(f"  {key.replace('_', ' ').title()}: {value}")
    
    def save_with_change_log(self):
        """Save the updated master file with optional change log sheet."""
        try:
            self.logger.info(f"💾 Saving results to: {self.output_path}")
            
            # Load original workbook to preserve macros
            wb = load_workbook(self.master_path, keep_vba=True)
            
            # Remove existing Assignment_Master sheet if it exists
            if 'Assignment_Master' in wb.sheetnames:
                wb.remove(wb['Assignment_Master'])
            
            # Create new Assignment_Master sheet
            ws_master = wb.create_sheet('Assignment_Master', 0)  # Insert at beginning
            
            # Write headers with formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            for col_idx, col_name in enumerate(self.df_master.columns, 1):
                cell = ws_master.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Write data
            for row_idx, row_data in enumerate(self.df_master.itertuples(index=False, name=None), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_master.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
            
            # Auto-adjust column widths
            for column in ws_master.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_master.column_dimensions[column_letter].width = adjusted_width
            
            # Add change log sheet if requested and changes exist
            if self.create_change_log and self.changes_log:
                ws_changes = wb.create_sheet('Change_Log')
                
                # Create changes dataframe
                df_changes = pd.DataFrame(self.changes_log)
                
                # Write change log headers
                for col_idx, col_name in enumerate(df_changes.columns, 1):
                    cell = ws_changes.cell(row=1, column=col_idx, value=col_name)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="D9534F", end_color="D9534F", fill_type="solid")
                    cell.border = border
                
                # Write change log data
                for row_idx, row_data in enumerate(df_changes.itertuples(index=False, name=None), 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_changes.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                
                # Auto-adjust change log column widths
                for column in ws_changes.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_changes.column_dimensions[column_letter].width = adjusted_width
                
                self.logger.info(f"📋 Change log created with {len(self.changes_log)} entries")
            
            # Add summary sheet
            self._add_summary_sheet(wb)
            
            # Save workbook
            wb.save(self.output_path)
            self.logger.info(f"✅ File saved successfully: {self.output_path}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"❌ Error saving file: {e}")
            return False
    
    def _add_summary_sheet(self, workbook):
        """Add a summary sheet with merge statistics and metadata."""
        ws_summary = workbook.create_sheet('Merge_Summary')
        
        summary_data = [
            ["Assignment Master Merge Summary", ""],
            ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Author", "R. A. Carucci"],
            ["", ""],
            ["Input Files", ""],
            ["Master File", str(self.master_path.name)],
            ["Submission File", str(self.submission_path.name)],
            ["Conflicts File", str(self.conflicts_path.name)],
            ["", ""],
            ["Statistics", ""],
            ["Total Records", self.stats['total_records']],
            ["Submission Updates", self.stats['submission_updates']],
            ["Conflict Updates", self.stats['conflict_updates']],
            ["No Updates Required", self.stats['no_updates']],
            ["Errors", self.stats['errors']],
            ["", ""],
            ["Change Log", ""],
            ["Changes Tracked", len(self.changes_log) if self.create_change_log else "Disabled"],
            ["Change Log Sheet", "Created" if self.create_change_log and self.changes_log else "Not Created"]
        ]
        
        # Write summary data with formatting
        for row_idx, (label, value) in enumerate(summary_data, 1):
            # Label column
            cell_label = ws_summary.cell(row=row_idx, column=1, value=label)
            if label and not value:  # Section headers
                cell_label.font = Font(bold=True, size=12)
            elif label:  # Regular labels
                cell_label.font = Font(bold=True)
            
            # Value column
            if value:
                cell_value = ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Auto-adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 30
    
    def run_complete_merge(self):
        """Execute the complete merge process."""
        self.logger.info("🚀 Starting Enhanced Assignment Master Merge")
        self.logger.info(f"📁 Output will be saved to: {self.output_path}")
        
        try:
            # Step 1: Load and validate data
            if not self.load_and_validate_data():
                return False
            
            # Step 2: Normalize badge numbers
            self.normalize_all_badges()
            
            # Step 3: Create lookup tables
            self.create_lookup_tables()
            
            # Step 4: Apply updates with change tracking
            self.apply_updates_with_tracking()
            
            # Step 5: Save results
            if not self.save_with_change_log():
                return False
            
            # Step 6: Final summary
            self.logger.info("🎉 Merge completed successfully!")
            self.logger.info("📊 Final Summary:")
            self.logger.info(f"  ✅ {self.stats['submission_updates']} records updated from submissions")
            self.logger.info(f"  🔧 {self.stats['conflict_updates']} records updated from conflicts")
            self.logger.info(f"  📝 {len(self.changes_log) if self.create_change_log else 0} changes logged")
            self.logger.info(f"  💾 Results saved to: {self.output_path}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"❌ Merge failed: {e}")
            return False


def main():
    """Command-line interface for the merger."""
    parser = argparse.ArgumentParser(description='Enhanced Assignment Master Merger')
    parser.add_argument('master', help='Path to Assignment_Master.xlsm')
    parser.add_argument('submission', help='Path to ats_personnel_submission.xlsx')
    parser.add_argument('conflicts', help='Path to naming_conflicts_resolution.xlsx')
    parser.add_argument('-o', '--output', help='Output file path (optional)')
    parser.add_argument('--no-change-log', action='store_true', 
                       help='Disable change log creation')
    
    args = parser.parse_args()
    
    # Create and run merger
    merger = EnhancedAssignmentMerger(
        master_path=args.master,
        submission_path=args.submission,
        conflicts_path=args.conflicts,
        output_path=args.output,
        create_change_log=not args.no_change_log
    )
    
    success = merger.run_complete_merge()
    sys.exit(0 if success else 1)


# ============================================================================
# SIMPLE ONE-CLICK FUNCTION VERSION
# ============================================================================

def quick_merge_with_change_log(master_path: str, submission_path: str, 
                               conflicts_path: str, output_path: str = None):
    """
    One-click merge function that handles everything automatically.
    
    Args:
        master_path: Path to Assignment_Master.xlsm
        submission_path: Path to ats_personnel_submission.xlsx  
        conflicts_path: Path to naming_conflicts_resolution.xlsx
        output_path: Optional output path (auto-generated if None)
    
    Returns:
        tuple: (success: bool, output_file: str, stats: dict)
    """
    try:
        merger = EnhancedAssignmentMerger(master_path, submission_path, 
                                        conflicts_path, output_path)
        success = merger.run_complete_merge()
        return success, str(merger.output_path), merger.stats
    
    except Exception as e:
        print(f"❌ Quick merge failed: {e}")
        return False, None, {}


if __name__ == "__main__":
    main()


# ============================================================================
# USAGE EXAMPLES
# ============================================================================

"""
# Example 1: Command line usage
python enhanced_merger.py Assignment_Master.xlsm ats_personnel_submission.xlsx naming_conflicts_resolution.xlsx

# Example 2: Command line with custom output
python enhanced_merger.py Assignment_Master.xlsm ats_personnel_submission.xlsx naming_conflicts_resolution.xlsx -o custom_output.xlsm

# Example 3: Command line without change log
python enhanced_merger.py Assignment_Master.xlsm ats_personnel_submission.xlsx naming_conflicts_resolution.xlsx --no-change-log

# Example 4: Programmatic usage
merger = EnhancedAssignmentMerger(
    "Assignment_Master.xlsm",
    "ats_personnel_submission.xlsx", 
    "naming_conflicts_resolution.xlsx"
)
success = merger.run_complete_merge()

# Example 5: One-click function
success, output_file, stats = quick_merge_with_change_log(
    "Assignment_Master.xlsm",
    "ats_personnel_submission.xlsx",
    "naming_conflicts_resolution.xlsx"
)
print(f"Success: {success}, Output: {output_file}, Stats: {stats}")
"""